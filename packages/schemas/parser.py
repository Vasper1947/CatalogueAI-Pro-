"""Parse BK bulk-upload templates (.xlsx) into TemplateSchema objects.

General by design: nothing is hardcoded to one template's field set. The
field-name row is detected, types come from file signals (data validations,
number formats), locked/formula columns from cell protection, and category /
rules from the Instructions sheet. A file that does not fit the expected shape
becomes a reported ParseFailure — never a guessed partial schema.
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from schemas.models import Field, ParseFailure, ParseReport, TemplateSchema

REQUIRED_SHEETS = ("Template", "Lookup", "Instructions")
_ID_RE = re.compile(r"([a-z_]+_id):\s*(cat_[A-Za-z0-9_-]+)")
_ROW_LIMIT_RE = re.compile(r"up to\s+([0-9,]+)\s+products", re.IGNORECASE)
_LOOKUP_REF_RE = re.compile(
    r"^'?(?P<sheet>[^'!]+)'?!\$?(?P<col>[A-Za-z]+)\$?\d+(?::\$?[A-Za-z]+\$?\d+)?$"
)
_KNOWN_HEADERS = {
    "general instructions",
    "template sections",
    "required fields",
    "product naming convention",
    "categorization ids (do not modify):",
}


class TemplateParseError(Exception):
    """Raised when a template cannot be parsed into a schema."""


def _detect_header_row(ws, scan: int = 6) -> int:
    """The field-name row = the densest row of text labels near the top."""
    best_row, best = 1, -1
    for r in range(1, scan + 1):
        count = sum(1 for c in ws[r] if isinstance(c.value, str) and c.value.strip())
        if count > best:
            best_row, best = r, count
    if best <= 1:
        raise TemplateParseError("could not locate a field-name row in the Template sheet")
    return best_row


def _clean_name(raw: str) -> tuple[str, bool, bool]:
    name = raw.strip()
    required = name.endswith("*")
    lowered = name.lower()
    conditional = "(cond)" in lowered
    name = name.rstrip("*").strip()
    name = re.sub(r"\((opt|cond)\)", "", name, flags=re.IGNORECASE).strip()
    return name, required, conditional


def _is_numeric_format(fmt: str | None) -> bool:
    if not fmt or fmt in ("General", "@"):
        return False
    if not any(ch in fmt for ch in "0#"):
        return False
    # Exclude date/time formats (which also carry digit placeholders in some locales).
    return not any(tok in fmt.lower() for tok in ("yy", "dd", "hh", "ss", "am/pm"))


def _validation_map(ws) -> dict[int, str]:
    """Column index -> list-validation formula1 string, for dropdown columns."""
    out: dict[int, str] = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or dv.formula1 is None:
            continue
        for rng in dv.sqref.ranges:
            for col in range(rng.min_col, rng.max_col + 1):
                out[col] = dv.formula1
    return out


def _read_lookups(ws) -> tuple[dict[str, list[str]], dict[str, tuple[str, list[str]]]]:
    """Lookup sheet -> ({name: values}, {column_letter: (name, values)})."""
    by_name: dict[str, list[str]] = {}
    by_col: dict[str, tuple[str, list[str]]] = {}
    if ws is None:
        return by_name, by_col
    for cell in ws[1]:
        if not (isinstance(cell.value, str) and cell.value.strip()):
            continue
        name = cell.value.strip()
        values: list[str] = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cell.column).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            values.append(str(v).strip())
        by_name[name] = values
        by_col[get_column_letter(cell.column)] = (name, values)
    return by_name, by_col


def _resolve_vocabulary(
    formula1: str, lookups_by_col: dict[str, tuple[str, list[str]]]
) -> tuple[str, list[str]]:
    f = (formula1 or "").strip()
    if f.startswith('"') and f.endswith('"'):
        inner = f[1:-1]
        return "inline", [v.strip() for v in inner.split(",") if v.strip()]
    m = _LOOKUP_REF_RE.match(f)
    if m and m.group("sheet") == "Lookup":
        name, values = lookups_by_col.get(m.group("col"), (None, []))
        return f"lookup:{name}", list(values)
    return "reference", []  # a validation we could not resolve to a vocabulary


def _section_map(ws, header_row: int) -> dict[int, str | None]:
    out: dict[int, str | None] = {}
    if header_row <= 1:
        return out
    current: str | None = None
    for cell in ws[header_row - 1]:
        if isinstance(cell.value, str) and cell.value.strip():
            current = cell.value.strip()
        out[cell.column] = current
    return out


def _block_after(lines: list[str], header: str) -> list[str]:
    try:
        start = next(i for i, ln in enumerate(lines) if ln.strip().lower() == header.lower())
    except StopIteration:
        return []
    block: list[str] = []
    for ln in lines[start + 1:]:
        if ln.strip().lower() in _KNOWN_HEADERS or ln.lower().startswith("pattern:"):
            break
        block.append(ln)
    return block


def _parse_instructions(ws) -> dict:
    lines = [
        v.strip()
        for row in ws.iter_rows(values_only=True)
        for v in row
        if isinstance(v, str) and v.strip()
    ]
    text = "\n".join(lines)
    ids = dict(_ID_RE.findall(text))
    # The breadcrumb is the line right after "generated for:" (works for a
    # 1-level category too, which has no ">" separator); fall back to any line
    # that clearly is a breadcrumb.
    breadcrumb = None
    for i, ln in enumerate(lines):
        if "generated for" in ln.lower():
            breadcrumb = lines[i + 1] if i + 1 < len(lines) else None
            break
    if breadcrumb is None:
        breadcrumb = next((ln for ln in lines if " > " in ln), None)
    naming = next((ln for ln in lines if ln.lower().startswith("pattern:")), None)
    row_limit = None
    m = _ROW_LIMIT_RE.search(text)
    if m:
        row_limit = int(m.group(1).replace(",", ""))
    media_options = [ln for ln in lines if "media" in ln.lower() or "url" in ln.lower()]
    required = [ln.lstrip("- ").strip() for ln in _block_after(lines, "Required Fields")]
    return {
        "breadcrumb": breadcrumb,
        "category_ids": ids,
        "naming_convention": naming,
        "row_limit": row_limit,
        "media_options": media_options,
        "required_fields": [r for r in required if r],
        "sections": _block_after(lines, "Template Sections"),
        "raw_lines": lines,
    }


def _path_from_filename(zip_category: str, filename: str) -> list[str]:
    stem = filename[:-5] if filename.lower().endswith(".xlsx") else filename
    parts = [p.replace("_", " ").strip() for p in stem.split("_-_")]
    return [zip_category, *parts]


def _zip_category_name(zip_path: str) -> str:
    stem = Path(zip_path).stem
    stem = stem.removeprefix("BK_")
    return stem.replace("_", " ").strip()


def parse_template(source, *, filename: str, zip_category: str) -> TemplateSchema:
    """Parse one .xlsx (path or file-like) into a TemplateSchema.

    Raises TemplateParseError if a required sheet is missing or the Template
    sheet has no detectable field-name row.
    """
    wb = openpyxl.load_workbook(source, data_only=False)
    try:
        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise TemplateParseError(f"missing required sheet(s): {missing}")

        tmpl = wb["Template"]
        header_row = _detect_header_row(tmpl)
        data_row = header_row + 1
        protected = bool(tmpl.protection.sheet)
        dv = _validation_map(tmpl)
        lookups_by_name, lookups_by_col = _read_lookups(wb["Lookup"])
        sections = _section_map(tmpl, header_row)

        fields: list[Field] = []
        for cell in tmpl[header_row]:
            if not (isinstance(cell.value, str) and cell.value.strip()):
                continue
            name, required, conditional = _clean_name(cell.value)
            col = cell.column
            body = tmpl.cell(row=data_row, column=col)
            locked = bool(body.protection.locked) if protected else False
            is_formula = body.data_type == "f"
            if col in dv:
                ftype = "dropdown"
                source_kind, vocab = _resolve_vocabulary(dv[col], lookups_by_col)
            elif _is_numeric_format(body.number_format):
                ftype, source_kind, vocab = "numeric", None, []
            else:
                ftype, source_kind, vocab = "string", None, []
            fields.append(
                Field(
                    name=name,
                    column=get_column_letter(col),
                    type=ftype,
                    required=required,
                    conditional=conditional,
                    locked=locked,
                    is_formula=is_formula,
                    section=sections.get(col),
                    dropdown_source=source_kind,
                    vocabulary=vocab,
                )
            )

        instructions = _parse_instructions(wb["Instructions"])
        if instructions["breadcrumb"]:
            category_path = [p.strip() for p in instructions["breadcrumb"].split(">")]
        else:
            category_path = _path_from_filename(zip_category, filename)

        return TemplateSchema(
            zip_category=zip_category,
            filename=filename,
            category_path=category_path,
            category_ids=instructions["category_ids"],
            fields=fields,
            lookups=lookups_by_name,
            instructions=instructions,
        )
    finally:
        wb.close()


def parse_zip(zip_path: str) -> tuple[list[TemplateSchema], list[ParseFailure]]:
    """Parse every .xlsx in a category ZIP. Per-file errors are captured, not raised."""
    zip_category = _zip_category_name(zip_path)
    schemas: list[TemplateSchema] = []
    failures: list[ParseFailure] = []
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx") or "__MACOSX" in name:
                continue
            base = name.rsplit("/", 1)[-1]
            try:
                data = zf.read(name)
                schema = parse_template(
                    io.BytesIO(data), filename=base, zip_category=zip_category
                )
                schemas.append(schema)
            except Exception as exc:  # noqa: BLE001
                # One malformed file must never abort the run or produce a guess.
                failures.append(
                    ParseFailure(
                        source=f"{Path(zip_path).name}::{base}",
                        reason=f"{type(exc).__name__}: {exc}",
                    )
                )
    return schemas, failures


def parse_all(folder_path: str) -> ParseReport:
    """Run over every ZIP in a folder, returning all schemas plus a failure report."""
    folder = Path(folder_path)
    all_schemas: list[TemplateSchema] = []
    all_failures: list[ParseFailure] = []
    found = 0
    for zip_path in sorted(folder.glob("*.zip")):
        with zipfile.ZipFile(zip_path) as zf:
            found += sum(
                1
                for n in zf.namelist()
                if n.lower().endswith(".xlsx") and "__MACOSX" not in n
            )
        schemas, failures = parse_zip(str(zip_path))
        all_schemas.extend(schemas)
        all_failures.extend(failures)
    return ParseReport(schemas=all_schemas, failures=all_failures, files_found=found)
