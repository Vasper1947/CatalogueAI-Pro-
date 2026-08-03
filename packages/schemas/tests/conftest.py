"""Synthetic .xlsx fixtures built with openpyxl — fast, deterministic, and
independent of the real template files changing.
"""

import openpyxl
import pytest
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _build_template(
    path,
    *,
    field_specs,
    breadcrumb,
    ids,
    lookup_cols=None,
    protect=True,
    section_row=False,
    naming="Pattern: [Brand]",
    row_limit=500,
    omit_sheet=None,
):
    wb = openpyxl.Workbook()
    tmpl = wb.active
    tmpl.title = "Template"
    header_row = 2 if section_row else 1
    data_row = header_row + 1
    if section_row:
        tmpl.cell(row=1, column=1, value="Product Information")

    for idx, spec in enumerate(field_specs, start=1):
        letter = get_column_letter(idx)
        label = spec["name"] + (" *" if spec.get("required") else "")
        tmpl.cell(row=header_row, column=idx, value=label)
        body = tmpl.cell(row=data_row, column=idx)
        kind = spec["kind"]
        if kind == "numeric":
            body.number_format = "#,##0.00"
        if kind == "formula":
            body.value = f"=A{data_row}&B{data_row}"
        body.protection = Protection(locked=spec.get("locked", kind == "formula"))
        if kind == "dropdown_inline":
            dv = DataValidation(type="list", formula1='"' + ",".join(spec["vocab"]) + '"')
            tmpl.add_data_validation(dv)
            dv.add(f"{letter}{data_row}:{letter}502")
        elif kind == "dropdown_lookup":
            col = spec["lookup_col"]
            dv = DataValidation(type="list", formula1=f"Lookup!${col}$2:${col}$100")
            tmpl.add_data_validation(dv)
            dv.add(f"{letter}{data_row}:{letter}502")

    if protect:
        tmpl.protection.sheet = True

    lookup = wb.create_sheet("Lookup")
    for ci, (cname, values) in enumerate((lookup_cols or {}).items(), start=1):
        lookup.cell(row=1, column=ci, value=cname)
        for ri, val in enumerate(values, start=2):
            lookup.cell(row=ri, column=ci, value=val)

    instr = wb.create_sheet("Instructions")
    lines = [
        "Bulk Upload Template Instructions",
        "This template is generated for:",
        breadcrumb,
        "Categorization IDs (do not modify):",
        *[f"{k}: {v}" for k, v in ids.items()],
        "General Instructions",
        "2. Fields marked with * are required. Fields marked (opt) are optional.",
        f"14. This template supports up to {row_limit} products.",
        "8. For media fields: place files in a media/ folder, ZIP, and upload the .zip.",
        "9. Alternatively, enter direct URLs for media files.",
        "Required Fields",
        *[f"- {s['name']}" for s in field_specs if s.get("required")],
        "Product Naming Convention",
        naming,
    ]
    for ri, val in enumerate(lines, start=1):
        instr.cell(row=ri, column=1, value=val)

    if omit_sheet:
        wb.remove(wb[omit_sheet])

    wb.save(str(path))
    wb.close()


@pytest.fixture
def simple_template(tmp_path):
    path = tmp_path / "Power_Tools.xlsx"
    _build_template(
        path,
        field_specs=[
            {"name": "Brand", "kind": "dropdown_inline", "required": True,
             "vocab": ["Acme", "Globex", "Initech"]},
            {"name": "Model", "kind": "string", "required": True},
            {"name": "Weight", "kind": "numeric"},
            {"name": "Color", "kind": "dropdown_lookup", "lookup_col": "A"},
        ],
        breadcrumb="Tools",
        ids={"category_id": "cat_tools1"},
        lookup_cols={"Color": ["Red", "Green", "Blue"]},
        naming="Pattern: [Brand] [Model]",
    )
    return path


@pytest.fixture
def deep_template(tmp_path):
    path = tmp_path / "deep.xlsx"
    _build_template(
        path,
        field_specs=[
            {"name": "Brand", "kind": "string", "required": True},
            {"name": "Length", "kind": "numeric", "required": True},
        ],
        breadcrumb="Building Materials > Steel & Reinforcements > High Yield Bars > 12mm",
        ids={"category_id": "cat_a", "subcategory_id": "cat_b",
             "sub_subcategory_id": "cat_c", "product_type_id": "cat_d"},
        section_row=True,
        naming="Pattern: [Brand] Bar (L)[Length]",
    )
    return path


@pytest.fixture
def locked_template(tmp_path):
    path = tmp_path / "locked.xlsx"
    _build_template(
        path,
        field_specs=[
            {"name": "Brand", "kind": "string", "required": True},
            {"name": "System Product Name", "kind": "formula", "locked": True},
            {"name": "Price", "kind": "numeric"},
        ],
        breadcrumb="Materials > Bricks",
        ids={"category_id": "cat_x"},
    )
    return path


@pytest.fixture
def malformed_template(tmp_path):
    path = tmp_path / "bad.xlsx"
    _build_template(
        path,
        field_specs=[{"name": "Brand", "kind": "string"}],
        breadcrumb="Materials",
        ids={"category_id": "cat_x"},
        omit_sheet="Instructions",
    )
    return path
