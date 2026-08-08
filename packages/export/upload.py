"""Package a written .xlsx into the ZIP shape BK's bulk-upload actually
accepts — re-read from a real parsed template's own Instructions sheet, not
from a summary of it:

    "8. For media fields: place files in a "media/" folder, ZIP together
    with this .xlsx file, and upload the .zip."
    "9. Alternatively, enter direct URLs for media files."

So: the .xlsx at the ZIP root, plus a media/ folder holding the actual
files. Every filename a Media-section column actually references is checked
against what's really being packaged — programmatically, not assumed — via
schemas.parser.parse_template (the same locked, already-proven parser
engine/writer.py's own self-verification uses) re-reading the xlsx that was
actually produced, not a cached copy of its schema.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from schemas.parser import _detect_header_row, parse_template

MEDIA_SECTION = "Media"
_URL_RE = re.compile(r"(?i)^https?://")


@dataclass
class PackageResult:
    output_path: str
    media_written: list[str] = field(default_factory=list)
    # Filenames a Media-section column references but that were NOT supplied
    # in media_files -- a real gap, reported, never silently assumed present.
    missing_media: list[str] = field(default_factory=list)
    # Filenames supplied in media_files but never referenced by any row --
    # not an error (extra captured images are kept, not dropped), but worth
    # surfacing so a human can see the file made it in without a home.
    unreferenced_media: list[str] = field(default_factory=list)


def _split_media_value(raw: str) -> list[str]:
    """A media cell may list one or more filenames, comma-separated per the
    template's own multi-select convention ("no spaces after comma"), or a
    direct URL (point 9) -- a URL is not a media/ filename to verify, since
    that file is never expected to live in the local media/ folder."""
    pieces = [p.strip() for p in raw.split(",") if p.strip()]
    return [p for p in pieces if not _URL_RE.match(p)]


def _referenced_media_filenames(xlsx_path: Path) -> set[str]:
    schema = parse_template(str(xlsx_path), filename=xlsx_path.name, zip_category="")
    media_fields = [f for f in schema.fields if f.section == MEDIA_SECTION]
    if not media_fields:
        return set()

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    try:
        ws = wb["Template"]
        data_row_start = _detect_header_row(ws) + 1
        referenced: set[str] = set()
        for f in media_fields:
            col = column_index_from_string(f.column)
            for row in range(data_row_start, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None or not str(value).strip():
                    continue
                referenced.update(_split_media_value(str(value)))
        return referenced
    finally:
        wb.close()


def package_for_upload(xlsx_path, media_files: dict[str, bytes], output_zip) -> PackageResult:
    """Build the upload-ready ZIP: xlsx_path at the root, media_files under
    media/. Verifies every real media-column filename in the xlsx against
    what is actually being packaged. Always writes the ZIP (a missing media
    file is a reported gap, not a reason to withhold otherwise-good data) —
    see PackageResult.missing_media / unreferenced_media for the real result.
    """
    xlsx_path = Path(xlsx_path)
    referenced = _referenced_media_filenames(xlsx_path)

    missing = sorted(name for name in referenced if name not in media_files)
    unreferenced = sorted(name for name in media_files if name not in referenced)

    output_zip = Path(output_zip)
    output_zip.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(xlsx_path, arcname=xlsx_path.name)
        for name, content in media_files.items():
            zf.writestr(f"media/{name}", content)

    return PackageResult(
        output_path=str(output_zip),
        media_written=sorted(media_files),
        missing_media=missing,
        unreferenced_media=unreferenced,
    )
