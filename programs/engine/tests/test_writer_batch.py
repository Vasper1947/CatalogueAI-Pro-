"""write_template_batch: N records -> one .xlsx, one row per record, sharing
headers/validations/protection -- same schema shape and self-verification
discipline as test_writer.py, generalized to multiple rows."""

import openpyxl
import pytest
from engine.populate import FieldResult, PopulationResult
from engine.writer import WriteVerificationError, write_template, write_template_batch


def _schema():
    return {
        "zip_category": "Test Category",
        "filename": "test.xlsx",
        "category_path": ["Test Category", "Widget"],
        "fields": [
            {
                "name": "Brand", "column": "A", "type": "dropdown", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "lookup:Brand",
                "vocabulary": [f"Brand{i}" for i in range(45)],
            },
            {
                "name": "Material", "column": "B", "type": "dropdown", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "inline",
                "vocabulary": ["Aluminum", "Stainless Steel", "PVC"],
            },
            {
                "name": "Length", "column": "C", "type": "numeric", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": None, "vocabulary": [],
            },
            {
                "name": "Length Unit", "column": "D", "type": "dropdown", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "inline", "vocabulary": ["m"],
            },
            {
                "name": "System Product Name", "column": "E", "type": "string",
                "required": False, "conditional": False, "locked": True, "is_formula": True,
                "section": "Naming", "dropdown_source": None, "vocabulary": [],
            },
            {
                "name": "Floor Price", "column": "F", "type": "numeric", "required": False,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Pricing", "dropdown_source": None, "vocabulary": [],
            },
        ],
        "lookups": {
            "Brand": [f"Brand{i}" for i in range(45)],
            "Material": ["Aluminum", "Stainless Steel", "PVC"],
            "Length Unit": ["m"],
        },
        "instructions": {
            "raw_lines": ["Generated for:", "Test Category > Widget", "up to 500 products"],
            "row_limit": 500,
        },
    }


def _population(overrides):
    fields = []
    for f in _schema()["fields"]:
        name = f["name"]
        if name in overrides and overrides[name] is not None:
            fields.append(
                FieldResult(name=name, required=f["required"], status="populated",
                             value=overrides[name])
            )
        else:
            fields.append(FieldResult(name=name, required=f["required"], status="needs_input"))
    return PopulationResult(
        category_path=["Test Category", "Widget"], fields=fields, status="incomplete",
        missing_required=[], populated_count=0, needs_input_count=0,
    )


def test_three_records_produce_three_rows_in_one_file(tmp_path):
    schema = _schema()
    pops = [
        _population({"Brand": "Brand1", "Material": "PVC", "Length": "1000 mm", "Length Unit": "m"}),
        _population({"Brand": "Brand2", "Material": "Aluminum", "Length": "2000 mm", "Length Unit": "m"}),
        _population({"Brand": "Brand3", "Material": "Stainless Steel", "Length": "3000 mm", "Length Unit": "m"}),
    ]
    out = tmp_path / "batch.xlsx"

    result = write_template_batch(pops, schema, out)

    assert len(result.row_results) == 3
    assert result.verification.structure_ok
    assert result.verification.value_mismatches == []

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Template"]
    # rows 3, 4, 5 -- one per record, DATA_ROW + offset
    assert ws.cell(row=3, column=1).value == "Brand1"
    assert ws.cell(row=4, column=1).value == "Brand2"
    assert ws.cell(row=5, column=1).value == "Brand3"
    assert ws.cell(row=3, column=3).value == 1.0  # Length column, metres
    assert ws.cell(row=4, column=3).value == 2.0
    assert ws.cell(row=5, column=3).value == 3.0


def test_headers_and_sections_written_once_not_per_row(tmp_path):
    schema = _schema()
    pops = [_population({"Brand": "Brand1"}), _population({"Brand": "Brand2"})]
    out = tmp_path / "batch.xlsx"

    write_template_batch(pops, schema, out)

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Template"]
    assert ws.cell(row=2, column=1).value == "Brand*"
    # No duplicate header text bled into row 4 (the second record's row).
    assert ws.cell(row=4, column=1).value == "Brand2"


def test_each_row_independently_respects_never_populate_and_locked(tmp_path):
    schema = _schema()
    pops = [
        _population({"Brand": "Brand1", "Floor Price": "999", "System Product Name": "X"}),
        _population({"Brand": "Brand2", "Floor Price": "111", "System Product Name": "Y"}),
    ]
    out = tmp_path / "batch.xlsx"

    result = write_template_batch(pops, schema, out)

    assert all("Floor Price" in r.blank_never_populate for r in result.row_results)
    assert all("System Product Name" not in r.written_fields for r in result.row_results)
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Template"]
    assert ws.cell(row=3, column=6).value is None  # Floor Price, row 1
    assert ws.cell(row=4, column=6).value is None  # Floor Price, row 2


def test_dropdown_validation_covers_every_row_not_just_the_first(tmp_path):
    schema = _schema()
    pops = [_population({"Material": "PVC"}), _population({"Material": "Aluminum"})]
    out = tmp_path / "batch.xlsx"

    write_template_batch(pops, schema, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Template"]
    material_dv = next(
        dv for dv in ws.data_validations.dataValidation
        if any(r.min_col == 2 for r in dv.sqref.ranges)
    )
    covered_rows = {r for rng in material_dv.sqref.ranges for r in range(rng.min_row, rng.max_row + 1)}
    assert {3, 4} <= covered_rows


def test_a_row_with_off_vocabulary_value_is_blank_others_still_write(tmp_path):
    schema = _schema()
    pops = [
        _population({"Material": "PVC"}),
        _population({"Material": "Titanium"}),  # not in vocabulary
    ]
    out = tmp_path / "batch.xlsx"

    result = write_template_batch(pops, schema, out)

    assert "Material" in result.row_results[0].written_fields
    assert "Material" in result.row_results[1].blank_invalid_dropdown


def test_write_template_single_record_matches_batch_of_one(tmp_path):
    # write_template() is now a thin wrapper over write_template_batch([...]).
    schema = _schema()
    pop = _population({"Brand": "Brand1", "Material": "PVC"})
    out = tmp_path / "single.xlsx"

    result = write_template(pop, schema, out)

    assert result.written_fields == ["Brand", "Material"]
    assert result.verification.structure_ok


def test_verification_failure_reports_which_row(tmp_path, monkeypatch):
    schema = _schema()
    pops = [_population({"Material": "PVC"}), _population({"Material": "Aluminum"})]
    out = tmp_path / "batch.xlsx"

    import engine.writer as writer_mod

    real_verify = writer_mod._verify

    def _tampering_verify(output_path, schema_, intended_fields, all_intended_values):
        wb = openpyxl.load_workbook(output_path)
        wb["Template"].cell(row=4, column=2, value="Tampered")  # corrupt row 2 (offset 1)
        wb.save(output_path)
        return real_verify(output_path, schema_, intended_fields, all_intended_values)

    monkeypatch.setattr(writer_mod, "_verify", _tampering_verify)

    with pytest.raises(WriteVerificationError, match="row 4"):
        write_template_batch(pops, schema, out)
