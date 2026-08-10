"""run_batch: detect + populate per staged record, grouped-by-category .xlsx
output, real row_limit splitting, and honest reporting of every non-matched
record -- never a silent drop."""

from engine.batch import DEFAULT_ROW_LIMIT, run_batch


def _schema(category_path, row_limit=None, required_field="Brand"):
    fields = [
        {"name": "Brand", "type": "dropdown", "required": True, "locked": False,
         "is_formula": False, "column": "A", "section": None,
         "dropdown_source": "inline", "vocabulary": ["Acme", "Zeta"]},
        {"name": "Grade", "type": "string", "required": False, "locked": False,
         "is_formula": False, "column": "B", "section": None,
         "dropdown_source": None, "vocabulary": []},
    ]
    instructions = {"raw_lines": [f"up to {row_limit or 500} products"]}
    if row_limit is not None:
        instructions["row_limit"] = row_limit
    return {
        "zip_category": "Cat", "filename": "cat.xlsx",
        "category_path": category_path, "fields": fields,
        "lookups": {"Brand": ["Acme", "Zeta"]}, "instructions": instructions,
    }


def _ev(brand="Acme", grade="A1"):
    rows = [{"field": "Brand", "value": brand, "absence": False}]
    if grade is not None:
        rows.append({"field": "Grade", "value": grade, "absence": False})
    return rows


def test_matched_records_grouped_by_category_write_one_file(tmp_path):
    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev()), ("p2", _ev(brand="Zeta"))]

    summary = run_batch(records, [schema], tmp_path)

    assert summary.total_records == 2
    assert len(summary.matched) == 2
    assert len(summary.written_files) == 1
    wf = summary.written_files[0]
    assert wf.record_ids == ["p1", "p2"]
    assert wf.category_path == ["Cat", "Widget"]
    assert summary.no_template_match == []
    assert summary.category_ambiguous == []
    assert summary.write_failures == []


def test_no_template_match_is_reported_not_dropped(tmp_path):
    schema = _schema(["Cat", "Widget"])
    # Evidence with fields that don't correspond to this schema at all.
    records = [("p1", [{"field": "Unrelated Thing", "value": "x", "absence": False}])]

    summary = run_batch(records, [schema], tmp_path)

    assert summary.total_records == 1
    assert len(summary.no_template_match) == 1
    assert summary.no_template_match[0].record_id == "p1"
    assert summary.no_template_match[0].reason is not None
    assert summary.matched == []
    assert summary.written_files == []


def test_row_limit_splits_into_multiple_files(tmp_path):
    schema = _schema(["Cat", "Widget"], row_limit=2)
    records = [(f"p{i}", _ev()) for i in range(5)]  # 5 records, limit 2 -> 3 files

    summary = run_batch(records, [schema], tmp_path)

    assert len(summary.matched) == 5
    assert len(summary.written_files) == 3
    sizes = sorted(len(wf.record_ids) for wf in summary.written_files)
    assert sizes == [1, 2, 2]
    all_ids = {rid for wf in summary.written_files for rid in wf.record_ids}
    assert all_ids == {f"p{i}" for i in range(5)}


def test_default_row_limit_used_when_schema_states_none(tmp_path):
    schema = _schema(["Cat", "Widget"], row_limit=None)  # helper omits the key entirely
    assert "row_limit" not in schema["instructions"]
    records = [("p1", _ev())]

    summary = run_batch(records, [schema], tmp_path)

    assert len(summary.written_files) == 1
    assert DEFAULT_ROW_LIMIT == 500  # documented fallback, not a silent guess


def test_genuinely_ambiguous_category_is_reported_and_excluded_from_output(tmp_path):
    # Two schemas with identical fields and no distinguishing evidence text --
    # a real, unresolved tie (see engine.detect.match_template), not a bug.
    # Must be reported, never silently resolved to one, never written.
    schema_a = _schema(["Cat", "Widget A"])
    schema_b = _schema(["Cat", "Widget B"])
    records = [("p1", _ev()), ("p2", _ev())]

    summary = run_batch(records, [schema_a, schema_b], tmp_path)

    assert len(summary.category_ambiguous) == 2
    assert {o.record_id for o in summary.category_ambiguous} == {"p1", "p2"}
    assert all(o.reason for o in summary.category_ambiguous)
    assert summary.matched == []
    assert summary.written_files == []


def test_output_files_are_real_and_readable(tmp_path):
    import openpyxl

    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev()), ("p2", _ev(brand="Zeta"))]

    summary = run_batch(records, [schema], tmp_path)

    wf = summary.written_files[0]
    wb = openpyxl.load_workbook(wf.output_path, data_only=True)
    ws = wb["Template"]
    assert ws.cell(row=3, column=1).value == "Acme"
    assert ws.cell(row=4, column=1).value == "Zeta"


def test_expand_variants_flag_off_by_default_leaves_variant_candidate_unexpanded(tmp_path):
    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev(brand="Acme/Zeta"))]  # 2 real whole-word matches -> variant_candidate

    summary = run_batch(records, [schema], tmp_path)

    assert len(summary.matched) == 1
    assert summary.variant_expansions == []
    wf = summary.written_files[0]
    assert len(wf.row_results) == 1  # not expanded -- one unresolved row
    assert "Brand" in wf.row_results[0].blank_variant_candidate


def test_expand_variants_flag_on_expands_into_real_rows(tmp_path):
    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev(brand="Acme/Zeta"))]

    summary = run_batch(records, [schema], tmp_path, expand_variants_flag=True)

    assert len(summary.matched) == 1
    assert len(summary.variant_expansions) == 1
    assert summary.variant_expansions[0].record_id == "p1"
    assert summary.variant_expansions[0].expanded_field == "Brand"
    assert summary.variant_expansions[0].option_count == 2
    wf = summary.written_files[0]
    assert len(wf.row_results) == 2  # one real row per option
    assert wf.record_ids == ["p1", "p1"]  # both rows trace back to the same source record
    assert all("Brand" in r.written_fields for r in wf.row_results)  # written, not blank

    import openpyxl
    wb = openpyxl.load_workbook(wf.output_path, data_only=True)
    ws = wb["Template"]
    assert {ws.cell(row=3, column=1).value, ws.cell(row=4, column=1).value} == {"Acme", "Zeta"}


def test_expand_variants_records_with_no_variant_fields_are_not_reported(tmp_path):
    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev())]  # clean single "Acme" value -- nothing to expand

    summary = run_batch(records, [schema], tmp_path, expand_variants_flag=True)

    assert summary.variant_expansions == []
    assert len(summary.written_files[0].row_results) == 1


def test_empty_records_list_produces_empty_summary(tmp_path):
    schema = _schema(["Cat", "Widget"])
    summary = run_batch([], [schema], tmp_path)

    assert summary.total_records == 0
    assert summary.matched == []
    assert summary.written_files == []


def test_forced_schema_override_skips_detection_entirely(tmp_path):
    schema = _schema(["Cat", "Widget A"])
    other = _schema(["Cat", "Widget B"])  # deliberately never passed to schemas=
    records = [("p1", _ev())]

    # With no forced_schema, [schema, other] would be a genuine ambiguous tie
    # (see test_genuinely_ambiguous_category_is_reported_and_excluded_from_output).
    # forced_schema bypasses match_template entirely -- a human's explicit,
    # not-a-guess override.
    summary = run_batch(records, [schema, other], tmp_path, forced_schema=schema)

    assert len(summary.matched) == 1
    assert summary.category_ambiguous == []
    assert len(summary.written_files) == 1
    assert summary.written_files[0].category_path == ["Cat", "Widget A"]


def test_a_chunk_write_failure_is_reported_not_raised(tmp_path, monkeypatch):
    # A self-verification failure for one chunk must not crash the whole
    # batch run or silently disappear -- it lands in write_failures.
    import engine.batch as batch_mod
    from engine.writer import WriteVerificationError

    def _always_fails(*_args, **_kwargs):
        raise WriteVerificationError("simulated mismatch")

    monkeypatch.setattr(batch_mod, "write_template_batch", _always_fails)

    schema = _schema(["Cat", "Widget"])
    records = [("p1", _ev())]

    summary = run_batch(records, [schema], tmp_path)

    assert len(summary.matched) == 1  # detection/population still succeeded
    assert summary.written_files == []
    assert len(summary.write_failures) == 1
    assert summary.write_failures[0].record_ids == ["p1"]
    assert "simulated mismatch" in summary.write_failures[0].error
