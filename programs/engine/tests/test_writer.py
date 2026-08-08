"""write_template tests: synthetic schema (mirrors the real Edge Trim schema's
shape — lookup-sourced Brand, inline dropdowns, a numeric field with a
companion "<Field> Unit" dropdown, a locked+formula field, a NEVER_POPULATE
field, and every real field sharing one of a few sections — the same shape
that exposed the header-row/section-row tie-detection bug during design)."""

import openpyxl
import pytest
from engine.populate import FieldResult, PopulationResult
from engine.writer import WriteVerificationError, _verify, write_template


def _schema():
    return {
        "zip_category": "Test Category",
        "filename": "test.xlsx",
        "category_path": ["Test Category", "Widget"],
        "fields": [
            {
                "name": "Brand", "column": "A", "type": "dropdown", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "lookup:Brand",
                "vocabulary": [f"Brand{i}" for i in range(45)],  # real Edge Trim scale
            },
            {
                "name": "Material", "column": "B", "type": "dropdown", "required": True,
                "conditional": True, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "inline",
                "vocabulary": ["Aluminum", "Stainless Steel", "PVC"],
            },
            {
                "name": "Length", "column": "C", "type": "numeric", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": None, "vocabulary": [],
            },
            {
                "name": "Length Unit", "column": "D", "type": "dropdown", "required": True,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": "inline", "vocabulary": ["m"],
            },
            {
                "name": "Notes", "column": "E", "type": "string", "required": False,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Info", "dropdown_source": None, "vocabulary": [],
            },
            {
                "name": "System Product Name", "column": "F", "type": "string",
                "required": False, "conditional": False, "locked": True, "is_formula": True,
                "section": "Naming", "dropdown_source": None, "vocabulary": [],
            },
            {
                "name": "Floor Price", "column": "G", "type": "numeric", "required": False,
                "conditional": False, "locked": False, "is_formula": False,
                "section": "Pricing", "dropdown_source": None, "vocabulary": [],
            },
        ],
        "lookups": {
            "Brand": [f"Brand{i}" for i in range(45)],
            "Material": ["Aluminum", "Stainless Steel", "PVC"],
            "Length Unit": ["m"],
        },
        "instructions": {
            "raw_lines": ["Generated for:", "Test Category > Widget", "up to 500 products"],
            "row_limit": 500,
        },
    }


def _population(overrides):
    """overrides: {field_name: value | None}. None -> needs_input."""
    fields = []
    for f in _schema()["fields"]:
        name = f["name"]
        if name in overrides and overrides[name] is not None:
            fields.append(
                FieldResult(name=name, required=f["required"], status="populated",
                             value=overrides[name])
            )
        else:
            fields.append(FieldResult(name=name, required=f["required"], status="needs_input"))
    return PopulationResult(
        category_path=["Test Category", "Widget"], fields=fields, status="incomplete",
        missing_required=[], populated_count=0, needs_input_count=0,
    )


def test_write_template_produces_a_verified_real_xlsx(tmp_path):
    schema = _schema()
    pop = _population({
        "Brand": "Brand12",
        "Material": "PVC",
        "Length": "2500 mm",  # canonical mm -> written in the field's real unit, metres
        "Length Unit": "m",
        # Notes left unpopulated -> needs_input
        # System Product Name: locked+formula, never written even if evidence existed
        # Floor Price: NEVER_POPULATE, never written even though a value could exist
        "Floor Price": "9999",
    })
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert out.exists()
    assert result.verification.structure_ok
    assert result.verification.value_mismatches == []
    assert result.verification.formula_fields_left_blank == ["System Product Name"]


def test_numeric_field_written_in_its_own_unit_not_canonical_mm(tmp_path):
    schema = _schema()
    pop = _population({"Length": "2500 mm", "Length Unit": "m"})
    out = tmp_path / "out.xlsx"

    write_template(pop, schema, out)

    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["Template"]
    assert ws.cell(row=3, column=3).value == 2.5  # column C = Length, metres not 2500mm


def test_dropdown_value_outside_vocabulary_is_left_blank(tmp_path):
    schema = _schema()
    pop = _population({"Material": "Aluminum Alloy"})  # real value, NOT in vocab
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert "Material" in result.blank_invalid_dropdown
    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb["Template"].cell(row=3, column=2).value is None


def test_dropdown_value_inside_vocabulary_is_written(tmp_path):
    schema = _schema()
    pop = _population({"Material": "PVC"})
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert "Material" in result.written_fields
    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb["Template"].cell(row=3, column=2).value == "PVC"


def test_floor_price_never_written_even_with_evidence(tmp_path):
    schema = _schema()
    pop = _population({"Floor Price": "12345"})
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert "Floor Price" in result.blank_never_populate
    assert "Floor Price" not in result.written_fields
    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb["Template"].cell(row=3, column=7).value is None  # column G = Floor Price

    # The column itself is still real: header present, matching the actual
    # template structure, unlike a fully omitted column.
    assert wb["Template"].cell(row=2, column=7).value == "Floor Price"


def test_locked_formula_field_never_written(tmp_path):
    schema = _schema()
    pop = _population({"System Product Name": "Should never appear"})
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert "System Product Name" not in result.written_fields
    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb["Template"].cell(row=3, column=6).value is None


def test_needs_input_field_tracked_and_blank(tmp_path):
    schema = _schema()
    pop = _population({})  # nothing populated
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert "Notes" in result.blank_needs_input


def test_large_lookup_sourced_dropdown_uses_lookup_sheet_reference(tmp_path):
    # Brand has 45 real values -- an inline "a,b,c,..." formula would exceed
    # Excel's ~255-character list-validation limit; the real template sources
    # it from the Lookup sheet instead, and write_template must match that.
    schema = _schema()
    pop = _population({"Brand": "Brand12"})
    out = tmp_path / "out.xlsx"

    write_template(pop, schema, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Template"]
    dvs = [dv for dv in ws.data_validations.dataValidation if dv.sqref.ranges]
    brand_dv = next(dv for dv in dvs if any(r.min_col == 1 for r in dv.sqref.ranges))
    assert brand_dv.formula1.startswith("Lookup!")


def test_required_and_conditional_field_round_trips_correctly(tmp_path):
    # Material is required=True and conditional=True together -- the exact
    # combination that exposed a header-text ordering bug during design
    # (schemas.parser._clean_name checks "ends with *" on the RAW header
    # BEFORE the "(cond)" marker is stripped, so "(cond)" must precede "*").
    schema = _schema()
    pop = _population({"Material": "PVC"})
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)

    assert result.verification.structure_ok  # would be False if required mis-parsed as False


def test_all_fields_sharing_sections_does_not_break_header_row_detection(tmp_path):
    # Every field in _schema() has a section (like the real Edge Trim schema)
    # -- if the section row were written densely enough to tie the header
    # row's string-cell count, schemas.parser._detect_header_row's "ties go
    # to the earlier row" rule would misdetect row 1 as the header row.
    schema = _schema()
    pop = _population({"Brand": "Brand1", "Material": "PVC", "Length": "1000 mm",
                        "Length Unit": "m"})
    out = tmp_path / "out.xlsx"

    result = write_template(pop, schema, out)  # raises WriteVerificationError if misdetected

    assert result.verification.structure_ok


def test_verify_reports_a_real_value_mismatch_when_the_file_diverges(tmp_path):
    # Directly proves the mandatory self-verification diff actually detects a
    # real divergence -- not just that it stays silent on a clean write.
    schema = _schema()
    pop = _population({"Material": "PVC"})
    out = tmp_path / "out.xlsx"
    write_template(pop, schema, out)

    # Corrupt the written file after the fact, independent of write_template,
    # to simulate the file no longer matching what was intended.
    wb = openpyxl.load_workbook(out)
    wb["Template"].cell(row=3, column=2, value="Something Else")
    wb.save(out)

    report = _verify(out, schema, schema["fields"], [{"Material": "PVC"}])

    assert report.value_mismatches == ["row 3 Material: wrote 'PVC', re-read 'Something Else'"]


def test_verify_reports_structure_failure_for_a_field_missing_from_the_file():
    schema = _schema()
    fake_fields = [*schema["fields"], {
        "name": "Nonexistent Field", "column": "Z", "type": "string", "required": False,
        "conditional": False, "locked": False, "is_formula": False, "section": None,
    }]
    # Build a minimal real file via write_template with the REAL field list,
    # then verify against fake_fields which claims one extra field exists.
    import tempfile
    from pathlib import Path
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "out.xlsx"
        write_template(_population({}), schema, out)
        report = _verify(out, schema, fake_fields, [{}])

    assert report.structure_ok is False


def test_write_verification_error_raised_is_never_silently_swallowed(tmp_path, monkeypatch):
    schema = _schema()
    pop = _population({"Material": "PVC"})
    out = tmp_path / "out.xlsx"

    # Force a genuine mismatch: after write_template's own save, but before
    # its self-verification runs, corrupt the file on disk.
    import engine.writer as writer_mod

    real_save_and_verify = writer_mod._verify

    def _tampering_verify(output_path, schema_, intended_fields, intended_values):
        wb = openpyxl.load_workbook(output_path)
        wb["Template"].cell(row=3, column=2, value="Tampered")
        wb.save(output_path)
        return real_save_and_verify(output_path, schema_, intended_fields, intended_values)

    monkeypatch.setattr(writer_mod, "_verify", _tampering_verify)

    with pytest.raises(WriteVerificationError):
        write_template(pop, schema, out)
