"""Write a matched, populated schema's data into a real .xlsx BK bulk-upload
template — reconstructed from the category's actual parsed template structure
(a TemplateSchema, as dict — see packages/schemas), not by editing an original
file, since no raw .xlsx files exist anywhere in this repo, only their parsed
structure (packages/schemas/data/*.json). NEVER_POPULATE/canonical come from
engine.detect — the SAME gate that keeps Floor Price out of scoring/population
keeps it out of the written file's VALUE too; there is no second, separate
exclusion list to drift out of sync. The Floor Price COLUMN itself is still
written (header, position, validation) exactly like every other real field —
only its value is withheld, matching "stays management's manual step": a
human still needs a normal, unlocked cell to type into, in the file's real
column position.

Self-verification is mandatory, not optional: every written file is
immediately re-parsed with packages/schemas' own, already-proven parser
(parse_template) and diffed against what write_template intended — field
names, types, required/locked flags, and column positions for the STRUCTURE;
the actual written cell values for the DATA, row by row. Any real mismatch
raises WriteVerificationError — a quietly-wrong file is never shipped.

One honest, documented, MECHANICAL limitation (not a convenience exclusion):
a formula column's original formula text (e.g. "System Product Name") is
never captured anywhere in the parsed schema — Field only records the
boolean is_formula, not the formula itself — so a reconstructed file cannot
contain a real, working formula for it, only its absence. This is not
something write_template chooses to skip checking; a blank cell mechanically
cannot re-parse as data_type == "f" no matter what is written, since there is
no formula text to write. is_formula is therefore excluded from the automated
structural diff for that reason alone, and is separately, explicitly reported
(never silently) — see VerificationReport.formula_fields_left_blank.

write_template() writes exactly one record (one data row) — the original,
still-supported entry point. write_template_batch() writes N records into the
SAME file as N data rows, sharing one set of column headers/validations/
protection, for the real case of many products in one BK category. Both
share the same per-row value logic and the same mandatory self-verification,
generalized to check every row, not just row 3.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from common.units import convert_from_mm, normalize_value
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from schemas.parser import parse_template

from engine.detect import NEVER_POPULATE, canonical
from engine.populate import PopulationResult

# Row 1: section labels (written only where a NEW section starts -- see
# _write_columns -- so its text-cell count stays below HEADER_ROW's for every
# real field). Row 2: field headers, every field, always. This ordering is
# what schemas.parser._section_map requires (it reads header_row - 1) and
# what _detect_header_row's own "most string cells wins, ties go to the
# earlier row" rule requires (see _write_columns's inline reasoning).
HEADER_ROW = 2
DATA_ROW = 3


class WriteVerificationError(Exception):
    """Raised when the mandatory self-verification re-parse/diff finds a real
    mismatch. The file has already been written to output_path, but must not
    be treated as trustworthy — callers should not ship it."""


@dataclass
class VerificationReport:
    structure_ok: bool
    value_mismatches: list[str] = field(default_factory=list)
    formula_fields_left_blank: list[str] = field(default_factory=list)


@dataclass
class RowWriteResult:
    """Per-record outcome within a written file — one per data row."""

    written_fields: list[str] = field(default_factory=list)
    blank_needs_input: list[str] = field(default_factory=list)
    blank_invalid_dropdown: list[str] = field(default_factory=list)
    blank_unresolved_numeric: list[str] = field(default_factory=list)
    blank_never_populate: list[str] = field(default_factory=list)
    blank_variant_candidate: list[str] = field(default_factory=list)


@dataclass
class WriteResult:
    """Single-record result — write_template()'s return value. Mirrors
    RowWriteResult's fields directly (it always describes exactly one row)."""

    output_path: str
    written_fields: list[str] = field(default_factory=list)
    blank_needs_input: list[str] = field(default_factory=list)
    blank_invalid_dropdown: list[str] = field(default_factory=list)
    blank_unresolved_numeric: list[str] = field(default_factory=list)
    blank_never_populate: list[str] = field(default_factory=list)
    blank_variant_candidate: list[str] = field(default_factory=list)
    verification: VerificationReport | None = None


@dataclass
class BatchWriteResult:
    """Multi-record result — write_template_batch()'s return value. One
    RowWriteResult per input PopulationResult, in the same order."""

    output_path: str
    row_results: list[RowWriteResult] = field(default_factory=list)
    verification: VerificationReport | None = None


def _header_text(f: dict) -> str:
    """Reconstruct the header cell text so schemas.parser._clean_name's exact
    detection order (required = raw.endswith("*"), checked BEFORE the
    "(cond)" parenthetical is stripped) round-trips correctly: the "(cond)"
    marker must come before a trailing "*", never after it."""
    name = f["name"]
    if f.get("conditional"):
        name = f"{name} (cond)"
    if f.get("required"):
        name = f"{name}*"
    return name


def _field_by_name(schema: dict, name: str) -> dict | None:
    return next((f for f in schema.get("fields", []) if f["name"] == name), None)


def _target_unit(schema: dict, base_name: str) -> str | None:
    """The real unit a numeric field is stored in, from its "<name> Unit"
    companion dropdown's single-entry vocabulary (e.g. Length -> "m") — a
    real, observed convention in this project's own parsed templates."""
    unit_field = _field_by_name(schema, f"{base_name} Unit")
    vocab = (unit_field or {}).get("vocabulary") or []
    return vocab[0] if len(vocab) == 1 else None


def _numeric_cell_value(schema: dict, field_name: str, raw_value: str) -> float | None:
    """raw_value (e.g. "2.5 m") -> a real number in the field's OWN schema
    unit -- never normalize_value's canonical mm blindly reused, and never a
    guessed number. None if it cannot be resolved cleanly."""
    normalized, source_unit, _confidence = normalize_value(raw_value)
    if not isinstance(normalized, (int, float)) or source_unit is None:
        return None
    target_unit = _target_unit(schema, field_name)
    if target_unit is None:
        return float(normalized)  # no companion unit field -- mm-canonical as-is
    return convert_from_mm(float(normalized), target_unit)


def _dropdown_formula(f: dict, lookup_col_letter: dict[str, str], num_rows: int) -> str:
    """A validation formula1 matching how THIS field's real vocabulary is
    sourced: a Lookup-sheet range reference when dropdown_source says so (the
    real convention for large vocabularies, e.g. a 45-brand list — Excel's
    inline list formula has a ~255-character limit an inline "a,b,c,..." for
    that many brands would exceed), inline otherwise. num_rows is unused here
    (the Lookup range is fixed by the vocabulary's own length regardless of
    how many data rows this file has) — kept as a parameter for symmetry with
    the caller's per-field setup and possible future per-row-count tuning."""
    del num_rows
    source = f.get("dropdown_source") or ""
    vocab = f.get("vocabulary") or []
    if source.startswith("lookup:"):
        col = lookup_col_letter.get(source[len("lookup:") :])
        if col:
            return f"Lookup!${col}$2:${col}${1 + len(vocab)}"
    return '"' + ",".join(vocab) + '"'


def _write_columns(
    tmpl, fields: list[dict], lookup_col_letter: dict[str, str]
) -> tuple[dict[str, int], dict[str, DataValidation]]:
    """One-time, row-count-independent setup: section labels, headers,
    dropdown validation objects, numeric formats. Returns ({field_name:
    column_index}, {field_name: DataValidation}) for the per-row writer to
    use — kept as local dicts, never stashed on the schema's own field dicts,
    since those may be a cached, shared object (e.g. engine/app.py's
    module-level schema cache) that must never carry transient per-write
    state between calls."""
    col_by_name: dict[str, int] = {}
    dv_by_name: dict[str, DataValidation] = {}
    prev_section = object()  # sentinel unequal to any real section value or None
    for f in fields:
        col = column_index_from_string(f["column"])
        col_by_name[f["name"]] = col
        section = f.get("section")
        if section and section != prev_section:
            tmpl.cell(row=1, column=col, value=section)
        prev_section = section

        tmpl.cell(row=HEADER_ROW, column=col, value=_header_text(f))

        if f.get("type") == "dropdown" and f.get("vocabulary"):
            dv = DataValidation(
                type="list", formula1=_dropdown_formula(f, lookup_col_letter, 0)
            )
            tmpl.add_data_validation(dv)
            dv_by_name[f["name"]] = dv
    return col_by_name, dv_by_name


def _write_row(
    tmpl, fields: list[dict], col_by_name: dict[str, int], dv_by_name: dict[str, DataValidation],
    row: int, schema: dict, population_result: PopulationResult,
) -> tuple[RowWriteResult, dict[str, object]]:
    """Write one record's values into row `row`. Returns (RowWriteResult,
    {field_name: value actually written}) — the latter feeds verification."""
    by_name = {fr.name: fr for fr in population_result.fields}
    result = RowWriteResult()
    intended_values: dict[str, object] = {}

    for f in fields:
        col = col_by_name[f["name"]]
        data_cell = tmpl.cell(row=row, column=col)

        if f.get("type") == "numeric":
            data_cell.number_format = "0.####"
        dv = dv_by_name.get(f["name"])
        if dv is not None:
            dv.add(data_cell)

        if f.get("locked") or f.get("is_formula"):
            # System/formula-managed -- never written; no formula text is
            # captured anywhere in the parsed schema to reconstruct one, and
            # a guessed value would be exactly the fabrication this whole
            # project exists to prevent. Left for the real system to auto-fill.
            continue
        data_cell.protection = Protection(locked=False)

        if canonical(f["name"]) in NEVER_POPULATE:
            result.blank_never_populate.append(f["name"])
            continue  # management's manual step -- never written regardless of evidence

        fr = by_name.get(f["name"])
        if fr is None or fr.value is None or fr.status != "populated":
            if fr is not None and fr.status == "variant_candidate":
                result.blank_variant_candidate.append(f["name"])
            else:
                result.blank_needs_input.append(f["name"])
            continue

        if f["type"] == "numeric":
            value = _numeric_cell_value(schema, f["name"], fr.value)
            if value is None:
                result.blank_unresolved_numeric.append(f["name"])
                continue
            data_cell.value = value
        elif f["type"] == "dropdown":
            vocab = f.get("vocabulary") or []
            if vocab and fr.value not in vocab:
                result.blank_invalid_dropdown.append(f["name"])
                continue
            data_cell.value = fr.value
        else:
            data_cell.value = fr.value
        intended_values[f["name"]] = data_cell.value
        result.written_fields.append(f["name"])

    return result, intended_values


def write_template(population_result: PopulationResult, schema: dict, output_path) -> WriteResult:
    """Write one populated record into a fresh .xlsx built from schema's real
    structure, then immediately re-parse and verify it. Raises
    WriteVerificationError on any real mismatch."""
    batch = write_template_batch([population_result], schema, output_path)
    row = batch.row_results[0]
    result = WriteResult(
        output_path=batch.output_path,
        written_fields=row.written_fields,
        blank_needs_input=row.blank_needs_input,
        blank_invalid_dropdown=row.blank_invalid_dropdown,
        blank_unresolved_numeric=row.blank_unresolved_numeric,
        blank_never_populate=row.blank_never_populate,
        blank_variant_candidate=row.blank_variant_candidate,
        verification=batch.verification,
    )
    return result


def write_template_batch(
    population_results: list[PopulationResult], schema: dict, output_path
) -> BatchWriteResult:
    """Write N populated records into ONE .xlsx, one data row each (starting
    at DATA_ROW), sharing one set of headers/validations/protection. Raises
    WriteVerificationError on any real mismatch across ANY row. Callers are
    responsible for the 500-row template limit (see engine/batch.py) — this
    function writes exactly len(population_results) rows, however many that
    is; it does not itself split or cap.
    """
    fields = schema.get("fields", [])

    wb = openpyxl.Workbook()
    tmpl = wb.active
    tmpl.title = "Template"
    tmpl.protection.sheet = True  # required for a written cell's locked=False to mean anything

    lookup_names = list((schema.get("lookups") or {}).keys())
    lookup_col_letter = {name: get_column_letter(i) for i, name in enumerate(lookup_names, start=1)}

    col_by_name, dv_by_name = _write_columns(tmpl, fields, lookup_col_letter)

    row_results: list[RowWriteResult] = []
    all_intended_values: list[dict[str, object]] = []
    for offset, population_result in enumerate(population_results):
        row_result, intended_values = _write_row(
            tmpl, fields, col_by_name, dv_by_name, DATA_ROW + offset, schema, population_result
        )
        row_results.append(row_result)
        all_intended_values.append(intended_values)

    _write_lookup_sheet(wb, schema)
    _write_instructions_sheet(wb, schema)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    verification = _verify(output_path, schema, fields, all_intended_values)
    if not verification.structure_ok or verification.value_mismatches:
        raise WriteVerificationError(
            f"{output_path}: structure_ok={verification.structure_ok}, "
            f"value_mismatches={verification.value_mismatches}"
        )
    return BatchWriteResult(
        output_path=str(output_path), row_results=row_results, verification=verification
    )


def _write_lookup_sheet(wb, schema: dict) -> None:
    ws = wb.create_sheet("Lookup")
    for i, (name, values) in enumerate((schema.get("lookups") or {}).items(), start=1):
        ws.cell(row=1, column=i, value=name)
        for r, v in enumerate(values, start=2):
            ws.cell(row=r, column=i, value=v)


def _write_instructions_sheet(wb, schema: dict) -> None:
    ws = wb.create_sheet("Instructions")
    # Replay the ORIGINAL template's own raw instruction lines verbatim --
    # reuses the exact text schemas.parser._parse_instructions already proved
    # it can correctly re-derive breadcrumb/category_ids/row_limit/required_
    # fields from, rather than reimplementing that formatting logic here.
    raw_lines = (schema.get("instructions") or {}).get("raw_lines") or []
    for i, line in enumerate(raw_lines, start=1):
        ws.cell(row=i, column=1, value=line)


def _verify(
    output_path, schema: dict, intended_fields: list[dict],
    all_intended_values: list[dict[str, object]],
) -> VerificationReport:
    """Re-parse the just-written file with the real, existing parser and diff
    it against what was intended -- structurally once, then by value for
    every row."""
    reparsed = parse_template(
        str(output_path),
        filename=schema.get("filename", "written.xlsx"),
        zip_category=schema.get("zip_category", ""),
    )
    reparsed_by_name = {f.name: f for f in reparsed.fields}

    formula_fields_left_blank: list[str] = []
    structure_ok = True
    for f in intended_fields:
        rf = reparsed_by_name.get(f["name"])
        if rf is None:
            structure_ok = False
            continue
        if f.get("is_formula"):
            # Mechanically expected: a blank cell can never re-parse as a
            # formula (no formula text was ever written -- see module
            # docstring). Recorded, not silently skipped.
            formula_fields_left_blank.append(f["name"])
            continue
        if (
            rf.type != f["type"]
            or bool(rf.required) != bool(f.get("required"))
            or bool(rf.locked) != bool(f.get("locked"))
            or rf.column != f["column"]
        ):
            structure_ok = False

    value_mismatches: list[str] = []
    wb = openpyxl.load_workbook(str(output_path), data_only=True)
    ws = wb["Template"]
    col_by_name = {f["name"]: f["column"] for f in intended_fields}
    for offset, intended_values in enumerate(all_intended_values):
        row = DATA_ROW + offset
        for name, expected in intended_values.items():
            col = column_index_from_string(col_by_name[name])
            actual = ws.cell(row=row, column=col).value
            if actual != expected:
                value_mismatches.append(
                    f"row {row} {name}: wrote {expected!r}, re-read {actual!r}"
                )
    wb.close()

    return VerificationReport(
        structure_ok=structure_ok,
        value_mismatches=value_mismatches,
        formula_fields_left_blank=formula_fields_left_blank,
    )
